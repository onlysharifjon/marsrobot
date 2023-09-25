import logging

import openpyxl
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.contrib.middlewares.logging import LoggingMiddleware
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import InputMediaPhoto

from defaullt.default import asosiy_menu
from defaullt.inline import language, who, important

# Your Telegram API token
TOKEN = '6514287083:AAF92CBSMQKpVXJM2gXwgSIfu7a5hceK5O0'
# Initialize the bot and dispatcher
bot = Bot(token=TOKEN,parse_mode='HTML')
dp = Dispatcher(bot, storage=MemoryStorage())
logging.basicConfig(level=logging.INFO)

# Middleware for logging
dp.middleware.setup(LoggingMiddleware())
DATABASE_DICT = {}


class Mars(StatesGroup):
    uzb_lang = State()
    modme = State()
    asosiy_men_state = State()
    space_shop = State()


# Echo handler
@dp.message_handler(commands=['start', 'help'])
async def send_welcome(message: types.Message):
    await message.answer(
        "Mars botiga xush kelibsiz! Iltimos, Til tanlang,\n\n–î–æ–±—Ä–æ –ø–æ–∂–∞–ª–æ–≤–∞—Ç—å –≤ Mars Bot! –ü–æ–∂–∞–ª—É–π—Å—Ç–∞, "
        "–≤—ã–±–µ—Ä–∏—Ç–µ —è–∑—ã–∫\n\nWelcome to Mars Bot! Please select a language",
        reply_markup=language)
    await Mars.uzb_lang.set()


@dp.callback_query_handler(text='Uzbek', state=Mars.uzb_lang)
async def uzb_l(call: types.CallbackQuery):
    await call.message.delete()
    await call.message.answer('Iltimos, kimligingizni ko`rsating))', reply_markup=who)


@dp.callback_query_handler(text='stud', state=Mars.uzb_lang)
async def student_login(call: types.CallbackQuery):
    await call.message.answer('Modme id ni kiriting: ')
    await call.message.delete()
    await Mars.modme.set()


p = 0


@dp.message_handler(content_types=types.ContentType.TEXT, state=Mars.modme)
async def texter(message: types.Message, state: FSMContext):
    id_student = message.text
    # print(id_student)
    # all values B2 in excel file
    wb = openpyxl.load_workbook('students.xlsx', 'rb')
    # activate workbook

    sheet = wb['Sheet']
    users = []

    for i in range(2, sheet.max_row + 1):
        users.append(sheet.cell(row=i, column=2).value)

    if int(id_student) in users:

        bolakaylar = []
        for k in range(2, sheet.max_row + 1):
            _ = []
            for p in range(1, 5):

                _.append(sheet.cell(row=k, column=p).value)
                if _ not in bolakaylar:
                    bolakaylar.append(_)

        async def adder():
            for t in bolakaylar:
                if str(t[1]) == str(id_student):
                    t.remove(0)
                    t.append(1)
            # for j in bolakaylar:
            #     print(j)

        count = 0
        for sim in bolakaylar:
            count += 1
            print(count)
            if sim[1] == int(id_student):
                if bolakaylar[count - 1][-1] == 1:
                    await message.answer('Bunday Akkaunt Oldin ro`yxatdan o`tgan')
                    await Mars.uzb_lang.set()
                elif bolakaylar[count - 1][-1] == 0:
                    await message.answer('Muvaffaqiyatli kirdingiz', reply_markup=asosiy_menu)
                    DATABASE_DICT[message.from_user.id] = int(id_student)
                    await state.finish()
                    await Mars.asosiy_men_state.set()
                    print(DATABASE_DICT)

                    await adder()

                else:
                    await message.answer('Tizimda qandaydir nosozlik!')

        # await state.finish()
        # break

        await state.finish()
        await Mars.asosiy_men_state.set()
    else:
        await message.answer('Bunday foydalanuvchi topilmadi!')


@dp.message_handler(text='üè´–û —à–∫–æ–ª–µ',state=Mars.asosiy_men_state)
async def profil(message:types.message,state:FSMContext):
    photo = open('defaullt/img.png', 'rb')
    await message.answer_photo(photo=photo)
    await message.answer('''
    üë®üèª‚Äçüíª Hamma kasb yaxshi, xavfsizi undan yaxshi!

Havola orqali o‚Äôting va videoni tomosha qiling:
https://www.instagram.com/p/CxYJKchiR9x/


ü•ä Boks sport turini sevuvchi har bir odam, hayotida bir marotaba bo'lsa ham professional bokschi bo'lishni orzu qilgan, lekin bu sport har birimizga ham to‚Äôg'ri kelmaydi. ¬´Mars IT School¬ª sizning farzandingiz hayotiga kamroq xavf tug‚Äôdiradigan va zamonaviy kasblarni qulay hamda qiziqarli holda taqdim etib, kelajakda o‚Äôz kasbining egasi bo'lishiga yordam beradi.

ü§ù Tanlovda adashmang, farzandingiz uchun hozirdan mustahkam poydevor quring!

<a href="https://t.me/Mars_yunusobod">üì© Yunusobod filiali</a>
<a href="https://t.me/Mars_tinchlik">üì© Tinchlik filiali</a>
<a href="https://t.me/Mars_Qutbiniso">üì© Chilonzor-Qutbiniso filialil;</a>
<a href="https://t.me/Mars_chilonzor18">üì© Chilonzor 18 filiali</a>

<b>üëáüèª Hoziroq izohlarda "+" belgisini qoldiring va biz siz bilan bog'lanamiz!</b>

<b>¬´Mars IT School¬ª ‚Äî bu kelajak!</b>

<b>üìû 78-777-77-57</b>
    ''')



@dp.message_handler(text='üí•Space shop',state=Mars.asosiy_men_state)
async def photo(message: types.Message):
    photos = [
        InputMediaPhoto(open('images/1photo.jpg', 'rb'),),
        InputMediaPhoto(open('images/2photo.jpg', 'rb')),
        InputMediaPhoto(open('images/3photo.jpg', 'rb')),
        InputMediaPhoto(open('images/4photo.jpg', 'rb')),
        InputMediaPhoto(open('images/5photo.jpg', 'rb')),
        InputMediaPhoto(open('images/6photo.jpg', 'rb')),
        InputMediaPhoto(open('images/7photo.jpg', 'rb')),
        InputMediaPhoto(open('images/8photo.jpg', 'rb')),
        InputMediaPhoto(open('images/9photo.jpg', 'rb')),
        InputMediaPhoto(open('images/10photo.jpg', 'rb')),
    ]
    await message.answer_media_group(media=photos)
    await message.answer('–í—ã–±–µ—Ä–∏—Ç–µ –ø—Ä–∏–∑',reply_markup=important)






if __name__ == '__main__':
    from aiogram import executor

    executor.start_polling(dp, skip_updates=True)
